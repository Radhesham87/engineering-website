"""Excel exports: Registered Users and Predictions."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT


def registered_users_xlsx(users: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registered Users"
    ws.append(["User ID", "Name", "Email", "Mobile", "City", "State",
               "Registration Date", "Approval Status", "Last Login",
               "Prediction Count"])
    for u in users:
        ws.append([
            u.id, u.name, u.email, u.mobile, u.city, u.state,
            u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else "",
            u.status.value if hasattr(u.status, "value") else str(u.status),
            u.last_login.strftime("%Y-%m-%d %H:%M") if u.last_login else "",
            len(u.predictions)])
    _style_header(ws)
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 45)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def predictions_xlsx(predictions: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    ws.append(["Prediction ID", "Date", "User", "Student Name", "Exam",
               "Mode", "Value", "Category", "Branches", "Districts",
               "Options"])
    for p in predictions:
        ws.append([
            p.id,
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            p.user.email if p.user else "",
            p.student_name, p.exam, p.mode, p.value, p.category,
            ", ".join(p.branches or []), ", ".join(p.districts or []),
            p.result_count])
    _style_header(ws)
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 45)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
